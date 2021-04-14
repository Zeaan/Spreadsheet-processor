import openpyxl as xl

def working_with_excel(name_of_excel_file,name_of_excel_sheet):
    wb = xl.load_workbook(name_of_excel_file)
    sheet = wb[name_of_excel_sheet]
    return sheet

#Make a function and start coding here (An example is given below)
def anyFunction(sheet):
    BCcount = 0
    Zcount = 0
    Bcount = 0
    others = 0

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,4)
        if cell.value == "Biochemistry":
            BCcount = BCcount + 1
        elif cell.value == "Zoology":
            Zcount = Zcount + 1
        elif cell.value == "Botany":
            Bcount = Bcount + 1
        else:
            others = others + 1

    AvgBC = (BCcount/(sheet.max_row-1)) * 100
    AvgB = (Bcount/(sheet.max_row-1)) * 100
    AvgZ = (Zcount/(sheet.max_row-1)) * 100
    AvgOthers = (others/(sheet.max_row-1)) * 100

    print("Percentage of Biochemistry students: "+str(AvgBC))
    print("Percentage of Botany students: "+str(AvgB))
    print("Percentage of Zoology students: "+str(AvgZ))
    print("Percentage of other students: "+str(AvgOthers))

name_of_excel_file = input("Enter the name of the Excel file ")
name_of_excel_sheet = input("Enter the name of the Excel sheet ")
sheet = working_with_excel(name_of_excel_file,name_of_excel_sheet)

anyFunction(sheet)